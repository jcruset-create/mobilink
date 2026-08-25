# -*- coding: utf-8 -*-
"""Libro de documentacion de sustitucion de tacografos.

Conserva literalmente los textos juridicos del libro original y anade lo que
exige la UNE 66102:2025 (informacion documentada, trazabilidad, anexo II del
RD 125/2017, registros de transferencia y destruccion, control documental).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import quote_sheetname, absolute_coordinate

OUT = "/home/user/mobilink/docs/plantillas/TACOGRAFOS_documentacion.xlsx"
TNR, ARIAL = "Times New Roman", "Arial"

f_titulo  = Font(name=ARIAL, size=14, bold=True)
f_seccion = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
f_label   = Font(name=ARIAL, size=10)
f_input   = Font(name=ARIAL, size=10, color="0000FF")
f_aviso   = Font(name=ARIAL, size=9, bold=True, color="C00000")
f_nota    = Font(name=ARIAL, size=9, italic=True, color="595959")
f_link    = Font(name=ARIAL, size=10, color="0563C1", underline="single")
d_txt     = Font(name=TNR, size=11)
d_txt_b   = Font(name=TNR, size=11, bold=True)
d_dato    = Font(name=TNR, size=11, bold=True)
d_tit     = Font(name=TNR, size=13, bold=True)
d_tit_j   = Font(name=TNR, size=12, bold=True)
d_peq     = Font(name=TNR, size=8, color="595959")
d_sec     = Font(name=TNR, size=10, bold=True, color="FFFFFF")

fill_sec  = PatternFill("solid", fgColor="1F4E79")
fill_sec2 = PatternFill("solid", fgColor="44546A")
fill_in   = PatternFill("solid", fgColor="FFF2CC")
fill_err  = PatternFill("solid", fgColor="FFC7CE")
fill_av   = PatternFill("solid", fgColor="FFF2F2")
fill_cab  = PatternFill("solid", fgColor="D9E2F3")

thin = Side(style="thin", color="BFBFBF")
box  = Border(left=thin, right=thin, top=thin, bottom=thin)
firma_l = Border(bottom=Side(style="thin", color="000000"))

wrap  = Alignment(wrap_text=True, vertical="top")
wrapj = Alignment(wrap_text=True, vertical="top", horizontal="justify")
ctr   = Alignment(horizontal="center", vertical="center")
ctrw  = Alignment(horizontal="center", vertical="center", wrap_text=True)
UNLOCK = Protection(locked=False)

wb = Workbook()
nombres = {}

# ============================================================ CONFIGURACION
cfg = wb.active
cfg.title = "CONFIGURACION"
cfg.sheet_view.showGridLines = False
cfg["A1"] = "CONFIGURACIÓN DEL CENTRO TÉCNICO"; cfg["A1"].font = f_titulo
cfg["A2"] = "Datos que no cambian. Se escriben aquí una vez; los documentos los toman de estas celdas."
cfg["A2"].font = f_nota

def cfg_seccion(r, texto):
    c = cfg.cell(r, 1, texto); c.font = f_seccion; c.fill = fill_sec
    for col in (2, 3): cfg.cell(r, col).fill = fill_sec
    return r + 1

def cfg_campo(r, name, label, valor, alto=False):
    cfg.cell(r, 1, label).font = f_label
    c = cfg.cell(r, 2, valor)
    c.font = f_input; c.fill = fill_in; c.border = box; c.protection = UNLOCK
    if alto: c.alignment = wrap
    nombres[name] = ("CONFIGURACION", f"B{r}")
    return r + 1

r = cfg_seccion(3, "IDENTIFICACIÓN DEL CENTRO")
for name, label, valor in [
    ("Cfg_Empresa", "Empresa", "COMERCIAL SEA S.A."),
    ("Cfg_CentroTecnico", "Centro técnico", "Centro técnico de Tacógrafos"),
    ("Cfg_NumCentro", "Contraseña / nº de centro", "E943009"),
    ("Cfg_Direccion1", "Dirección (línea 1)", "Pol.Ind. Riu Clar"),
    ("Cfg_Direccion2", "Dirección (línea 2)", "C/ Coure, 27"),
    ("Cfg_Ciudad", "Código postal y ciudad", "43006 Tarragona"),
    ("Cfg_CiudadFirma", "Ciudad de firma", "Tarragona"),
    ("Cfg_Email", "Email del centro", "jcruset@comercialsea.es"),
    ("Cfg_Destinatario", "Destinatario administración", "Direcció General de Transports i Mobilitat"),
]:
    r = cfg_campo(r, name, label, valor)

r += 1
r = cfg_seccion(r, "RESPONSABLES (UNE 66102:2025, 7.5.2 y anexo C)")
for name, label, valor in [
    ("Cfg_ResponsableTecnico", "Responsable técnico", "Jordi Cruset"),
    ("Cfg_Elaborado", "Documentación elaborada por", "Jordi Cruset"),
    ("Cfg_Aprobado", "Documentación aprobada por", "Dirección"),
]:
    r = cfg_campo(r, name, label, valor)

r += 1
r = cfg_seccion(r, "CONTROL DOCUMENTAL (UNE 66102:2025, 7.5.2 y 7.5.3)")
for name, label, valor in [
    ("Cfg_Version", "Versión del formato", "02"),
    ("Cfg_FechaEdicion", "Fecha de edición del formato", "2026-08-22"),
    ("Cfg_CodJust", "Código formato — justificante transferencia", "CTT-F-01"),
    ("Cfg_CodCliente", "Código formato — acuse intransferibilidad", "CTT-F-03"),
    ("Cfg_CodAdmin", "Código formato — comunicación administración", "CTT-F-04"),
    ("Cfg_CodRegistro", "Código formato — registro de transferencias", "CTT-R-01"),
    ("Cfg_Conservacion", "Conservación de los archivos transferidos", "1 año desde la transferencia (RD 125/2017, nota F)"),
]:
    r = cfg_campo(r, name, label, valor)
cfg[f"B{r-7}"].number_format = "DD/MM/YYYY"

r += 1
r = cfg_seccion(r, "ENLACES")
for name, label, valor in [
    ("Cfg_UrlTramite", "Trámite Generalitat (petició genèrica)",
     "https://web.gencat.cat/ca/tramits/tramits-temes/Peticio-generica"),
    ("Cfg_UrlTramiteOVT", "Trámite Generalitat (formulario OVT)",
     "https://ovt.gencat.cat/gsitgf/AppJava/traint/renderitzar.do?reqCode=inicial&set-locale=ca_ES"
     "&idioma=ca_ES&idServei=ING001SOLC&urlRetorn=https%3A%2F%2Fweb.gencat.cat%2Fca%2Ftramits%2F"
     "tramits-temes%2FPeticio-generica%3Fmoda%3D1&tpst=ae56fd5821cffc7f4784ac8db0c0e1a0"),
]:
    r = cfg_campo(r, name, label, valor, alto=True)

r += 1
r = cfg_seccion(r, "LISTAS")
lst_op_ini = r
for v in ["Transferencia correcta", "Intransferibilidad"]: cfg.cell(r, 1, v).font = f_label; r += 1
lst_op_fin = r - 1
lst_mod_ini = r
for v in ["En mano", "Email", "Mensajería", "Correo certificado"]: cfg.cell(r, 2, v).font = f_label; r += 1
lst_mod_fin = r - 1
lst_sn_ini = r
for v in ["Sí", "No"]: cfg.cell(r, 3, v).font = f_label; r += 1
lst_sn_fin = r - 1

r += 1
r = cfg_seccion(r, "FRAGMENTOS DE TEXTO LEGAL (editar sólo ante cambio normativo)")
TXT = [
 ("Txt_J_Informado", "He sido informado por el centro técnico de tacógrafos de "),
 ("Txt_J_Contrasena", "con contraseña "),
 ("Txt_J_Marca", ", de que el tacógrafo marca:"),
 ("Txt_En", "En "), ("Txt_A", ", a "), ("Txt_Sello", "Sello "),
 ("Txt_Entrega_Si", "***Se entrega tacógrafo Averiado"),
 ("Txt_Achatarrar_Si", "***El tacógrafo se achatarrará"),
 ("Txt_Fitxer", " Certificat de Intransferibilitat"),
 ("Txt_Confid", "Los datos contenidos en la memoria de la unidad intravehicular tienen carácter "
                "confidencial. El solicitante debe evaluar la confidencialidad de los datos transferidos "
                "para el procedimiento de remisión que elija. El centro técnico no será responsable de la "
                "violación de la confidencialidad de los datos durante su remisión (nota E del anexo II del "
                "Real decreto 125/2017)."),
 ("Txt_Custodia", "Los datos recuperados de la unidad instalada en el vehículo se guardarán en el centro "
                  "técnico durante un año desde la fecha de la transferencia. Una vez cumplido dicho plazo, "
                  "los datos serán destruidos (nota F del anexo II del Real decreto 125/2017)."),
 ("Txt_Titularidad", "Se ha presentado al centro técnico documento que avala la titularidad de los datos "
                     "por parte de la empresa de transportes, verificado y archivado por el centro (nota D "
                     "del anexo II del Real decreto 125/2017)."),
 ("Txt_Cat_1", "En compliment del requisit expressat en la disposició addicional primera, apartat 10, del "
               "Reial decret 125/2017, els remetem còpia del certificat de intransferibilitat corresponent "
               "a la substitució del tacògraf: Model: "),
 ("Txt_Cat_2", ", Nº de Sèrie: "), ("Txt_Cat_3", ", muntat en el vehicle: "),
 ("Txt_Cat_4", ", Nº d'informe / certificat: "), ("Txt_Cat_5", ", Data Informe: "),
 ("Txt_Cat_6", ". Sol·licito que es doni el tràmit acceptat."),
]
for name, valor in TXT:
    cfg.cell(r, 1, name).font = f_nota
    c = cfg.cell(r, 2, valor); c.font = f_label; c.alignment = wrap
    nombres[name] = ("CONFIGURACION", f"B{r}")
    r += 1

cfg.column_dimensions["A"].width = 44
cfg.column_dimensions["B"].width = 82
cfg.column_dimensions["C"].width = 14
cfg.protection.sheet = True

# ============================================================ DATOS
dat = wb.create_sheet("DATOS")
dat.sheet_view.showGridLines = False
dat["A1"] = "DATOS DE LA INTERVENCIÓN"; dat["A1"].font = f_titulo
dat["A2"] = "Rellena sólo las celdas amarillas. Todos los documentos se generan desde aquí."
dat["A2"].font = f_nota

# tipo: in / date / list / listm / listsn ; req: always / transf / intr / no
CAMPOS = [
 ("sec", "CLIENTE / EMPRESA DE TRANSPORTES", None, None, None),
 ("in",  "Empresa",                        "Dat_Empresa",        "COMERCIAL TANK FOODS S.L.", "always"),
 ("in",  "Nombre de quien autoriza",       "Dat_Nombre",         "Joan Pla Serra",   "always"),
 ("in",  "DNI / NIF de quien autoriza",    "Dat_Nif",            "39887654T",        "always"),
 ("listsn","Documento de titularidad aportado", "Dat_DocTitularidad", "Sí",           "always"),
 ("sec", "VEHÍCULO", None, None, None),
 ("in",  "Matrícula",                      "Dat_Matricula",      "7567MPF",          "always"),
 ("in",  "Nº de bastidor",                 "Dat_Bastidor",       "VF3XXXXXXXXXXXXXX","no"),
 ("sec", "TACÓGRAFO SUSTITUIDO (unidad intravehicular)", None, None, None),
 ("in",  "Marca / fabricante",             "Dat_Marca",          "VDO",              "always"),
 ("in",  "Modelo de la unidad",            "Dat_Modelo",         "1381.7550303006",  "always"),
 ("in",  "Nº de serie",                    "Dat_Serie",          "1000567",          "always"),
 ("sec", "INTERVENCIÓN", None, None, None),
 ("in",  "Nº informe / certificado",       "Dat_NumInforme",     "E943009003781L",   "always"),
 ("date","Fecha informe",                  "Dat_FechaInforme",   "2025-03-10",       "always"),
 ("date","Fecha entrega al cliente",       "Dat_FechaEntrega",   "2025-03-14",       "intr"),
 ("list","Tipo de operación",              "Dat_TipoOperacion",  "Intransferibilidad","always"),
 ("in",  "Técnico que interviene",         "Dat_Tecnico",        "Marc Roig",        "always"),
 ("sec", "TRANSFERENCIA DE DATOS", None, None, None),
 ("date","Fecha de transferencia",         "Dat_FechaTransf",    "2025-03-10",       "transf"),
 ("date","Fecha de envío",                 "Dat_FechaEnvio",     "2025-03-11",       "transf"),
 ("sec", "ENTREGA DE LOS DATOS DESCARGADOS", None, None, None),
 ("listm","Modalidad de entrega",          "Dat_Modalidad",      "En mano",          "transf"),
 ("sec", "PERSONA QUE RECIBE EL CERTIFICADO (acuse de intransferibilidad)", None, None, None),
 ("in",  "Nombre de la persona receptora", "Dat_NombreReceptor", "Marta Solé Vidal", "intr"),
 ("in",  "DNI de la persona receptora",    "Dat_DniReceptor",    "40123456X",        "intr"),
 ("sec", "TACÓGRAFO AVERIADO", None, None, None),
 ("listsn","Se entrega al cliente",        "Dat_EntregaAparato", "No",               "intr"),
 ("calc", "Se achatarrará (excluyente)",   "Dat_Achatarrar",     None,               "no"),
]

dv_op  = DataValidation(type="list", formula1=f"=CONFIGURACION!$A${lst_op_ini}:$A${lst_op_fin}", allow_blank=True)
dv_mod = DataValidation(type="list", formula1=f"=CONFIGURACION!$B${lst_mod_ini}:$B${lst_mod_fin}", allow_blank=True)
dv_sn  = DataValidation(type="list", formula1=f"=CONFIGURACION!$C${lst_sn_ini}:$C${lst_sn_fin}", allow_blank=True)
dv_fecha = DataValidation(type="date", operator="between", formula1="DATE(2000,1,1)",
                          formula2="DATE(2100,12,31)", allow_blank=True,
                          errorTitle="Fecha no válida", showErrorMessage=True,
                          error="Introduce una fecha real en formato dd/mm/aaaa.")
for dv in (dv_op, dv_mod, dv_sn, dv_fecha): dat.add_data_validation(dv)

r = 4
filas = []
for kind, label, name, ejemplo, req in CAMPOS:
    if kind == "sec":
        c = dat.cell(r, 1, label); c.font = f_seccion; c.fill = fill_sec
        for col in (2, 3): dat.cell(r, col).fill = fill_sec
        r += 1
        continue
    dat.cell(r, 1, label).font = f_label
    if kind == "calc":
        # Excluyente por construccion: es lo contrario de "se entrega al cliente",
        # igual que en el libro original, donde un unico SI/NO decidia las dos frases.
        c = dat.cell(r, 2, f'=IF(B{r-1}="Sí","No",IF(B{r-1}="No","Sí",""))')
        c.font = Font(name=ARIAL, size=10, bold=True); c.border = box
        dat.cell(r, 3, "automático").font = f_nota
        dat.cell(r, 4, 0)
        nombres[name] = ("DATOS", f"B{r}")
        r += 1
        continue
    c = dat.cell(r, 2, ejemplo)
    c.font = f_input; c.fill = fill_in; c.border = box; c.protection = UNLOCK
    if kind == "date":
        c.number_format = "DD/MM/YYYY"; dv_fecha.add(c)
    if kind == "list": dv_op.add(c)
    if kind == "listm": dv_mod.add(c)
    if kind == "listsn": dv_sn.add(c)
    if name == "Dat_Matricula":
        c.font = Font(name=ARIAL, size=10, color="0000FF", bold=True)
    nombres[name] = ("DATOS", f"B{r}")
    filas.append((r, req))
    r += 1
ultima = r - 1
fila_tipo = nombres["Dat_TipoOperacion"][1][1:]

for fila, req in filas:
    if req == "no":
        f = ""
    elif req == "always":
        f = f'=IF(B{fila}="","⚠ Obligatorio","")'
    elif req == "transf":
        f = f'=IF(AND($B${fila_tipo}="Transferencia correcta",B{fila}=""),"⚠ Obligatorio","")'
    else:
        f = f'=IF(AND($B${fila_tipo}="Intransferibilidad",B{fila}=""),"⚠ Obligatorio","")'
    if f:
        c = dat.cell(fila, 3, f); c.font = f_aviso; c.fill = fill_av
        dat.cell(fila, 4, f'=IF(C{fila}<>"",1,0)').font = f_nota
    else:
        dat.cell(fila, 3, "opcional").font = f_nota
        dat.cell(fila, 4, 0)

prim = filas[0][0]
dat["B2"] = (f'=IF(SUM(D{prim}:D{ultima})=0,"✔ Datos completos",'
             f'"⚠ Faltan "&SUM(D{prim}:D{ultima})&" campo(s) obligatorio(s)")')
dat["B2"].font = Font(name=ARIAL, size=11, bold=True)
dat.conditional_formatting.add("B2", FormulaRule(formula=[f'SUM(D{prim}:D{ultima})>0'], fill=fill_err))
dat.conditional_formatting.add(f"B{prim}:B{ultima}", FormulaRule(formula=[f'$C{prim}="⚠ Obligatorio"'], fill=fill_err))

dat.cell(ultima + 2, 1, "Destrucción de los archivos").font = f_label
dat.cell(ultima + 2, 2,
    f'=IF({nombres["Dat_FechaTransf"][1]}="","",'
    f'"Conservar hasta "&RIGHT("0"&DAY(EDATE({nombres["Dat_FechaTransf"][1]},12)),2)&"/"&'
    f'RIGHT("0"&MONTH(EDATE({nombres["Dat_FechaTransf"][1]},12)),2)&"/"&'
    f'YEAR(EDATE({nombres["Dat_FechaTransf"][1]},12))&" y destruir después (nota F, RD 125/2017)")').font = f_nota

dat.column_dimensions["A"].width = 38
dat.column_dimensions["B"].width = 42
dat.column_dimensions["C"].width = 16
dat.column_dimensions["D"].hidden = True
dat.protection.sheet = True

# ============================================================ helpers
def ref(name):
    s, c = nombres[name]
    return f"{quote_sheetname(s)}!{absolute_coordinate(c)}"

def txt(name):
    return f'IF({ref(name)}="","",{ref(name)})'

def fecha_txt(name):
    d = ref(name)
    return (f'IF({d}="","",RIGHT("0"&DAY({d}),2)&"/"&RIGHT("0"&MONTH({d}),2)&"/"&YEAR({d}))')

def setup_a4(ws, area):
    ws.sheet_view.showGridLines = False
    ws.print_area = area
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.7874
    ws.page_margins.top = ws.page_margins.bottom = 0.5906
    ws.oddFooter.right.text = "Página &P de &N"
    ws.oddFooter.right.size = 8

def cab_centro(ws, fila=1):
    for i, name in enumerate(["Cfg_Empresa", "Cfg_CentroTecnico", "Cfg_NumCentro",
                              "Cfg_Direccion1", "Cfg_Direccion2", "Cfg_Ciudad"]):
        c = ws.cell(fila + i, 1, f"={ref(name)}")
        c.font = d_txt_b if i == 0 else d_txt

def pie_control(ws, fila, cod, ancho="G"):
    """Identificacion del documento — UNE 66102:2025, 7.5.2."""
    ws.merge_cells(f"A{fila}:{ancho}{fila}")
    c = ws.cell(fila, 1,
        f'="Formato "&{ref(cod)}&" · Versión "&{ref("Cfg_Version")}&" · Edición "&'
        f'{fecha_txt("Cfg_FechaEdicion")}&" · Elaborado: "&{ref("Cfg_Elaborado")}&'
        f'" · Aprobado: "&{ref("Cfg_Aprobado")}&" · UNE 66102:2025"')
    c.font = d_peq; c.alignment = ctr
    c.border = Border(top=Side(style="thin", color="000000"))

# ============================================================ JUSTIFICANTE
j = wb.create_sheet("JUSTIFICANTE TRANSFERENCIA")
j["B1"] = f'={ref("Cfg_Empresa")}'; j["B1"].font = d_txt_b
j.merge_cells("B2:H2")
j["B2"] = ("JUSTIFICANTE PARA AUTORIZAR TODO EL PROCESO DE TRANSFERENCIA, EN EL CAMBIO DE UN "
           "TACÓGRAFO DIGITAL CON DESCARGA DE LA MEMORIA INTERNA DEL MISMO.")
j["B2"].font = d_tit_j; j["B2"].alignment = ctrw
j.row_dimensions[2].height = 46
for fila, f in [
    (3, f'="Yo, "&{txt("Dat_Nombre")}'),
    (4, f'="con N.I.F. nº: "&{txt("Dat_Nif")}'),
    (5, f'="en representación de la empresa de transportes: "&{txt("Dat_Empresa")}'),
    (6, f'="propietaria del vehículo matrícula: "&UPPER({txt("Dat_Matricula")})'),
    (7, f'={ref("Txt_J_Informado")}&{ref("Cfg_Empresa")}&","'),
    (8, f'={ref("Txt_J_Contrasena")}&{ref("Cfg_NumCentro")}&{ref("Txt_J_Marca")}&" "&{txt("Dat_Marca")}'),
    (9, f'="modelo: "&{txt("Dat_Modelo")}&"   y número de serie: "&{txt("Dat_Serie")}'),
]:
    j.merge_cells(f"B{fila}:H{fila}")
    j.cell(fila, 2, f).font = d_txt

TEXTOS_J = {
 10: ("instalado en el vehículo de matrícula indicada arriba, que ha de cambiarse, puede ser "
      "correctamente descargado en el contenido de su memoria interna."),
 11: ("Siendo representante de la empresa propietaria del vehículo o estando autorizado por la "
      "Dirección de la misma, doy la indicación y autorizo, al centro técnico de tacógrafos de "
      "COMERCIAL SEA, S.A. a la descarga de la mencionada memoria interna del tacógrafo."),
 12: ("Quedo advertido, en caso de no ser uno de los representantes de dicha empresa, de que debo "
      "autorizar todo este proceso habiendo recibido indicación expresa de la Dirección de la misma a "
      "actuar en el sentido en el que me manifiesto y que el proceso no tendrá respaldo legal si yo "
      "falseo esta autorización."),
 13: ("Doy la indicación, así mismo, de que los archivos de transferencia originados en la mencionada "
      "descarga sean entregados a la empresa propietaria del vehículo a través del medio indicado "
      "debajo, de entre las cuatro posibles opciones indicadas en el punto 6 de la disposición "
      "adicional primera del Real decreto 125/2017 (señálese la opción que proceda)"),
}
for fila, t in TEXTOS_J.items():
    j.merge_cells(f"B{fila}:H{fila}")
    c = j.cell(fila, 2, t); c.font = d_txt; c.alignment = wrapj
for fila, alto in [(10, 32), (11, 46), (12, 60), (13, 60)]:
    j.row_dimensions[fila].height = alto

for fila, t, clave in [
 (14, "Entrega en mano a una persona designada y autorizada para tal por la empresa propietaria del "
      "vehículo de matrícula arriba indicada", "En mano"),
 (15, "Entrega por medios electrónicos (por ejemplo, email)", "Email"),
 (16, "Entrega a través de empresa de mensajería", "Mensajería"),
 (17, "Entrega por correo certificado", "Correo certificado"),
]:
    j.merge_cells(f"B{fila}:G{fila}")
    c = j.cell(fila, 2, t); c.font = d_txt; c.alignment = wrap; c.border = box
    m = j.cell(fila, 8, f'=IF({ref("Dat_Modalidad")}="{clave}","X","")')
    m.font = d_dato; m.alignment = ctr; m.border = box
j.row_dimensions[14].height = 30

# clausulas UNE / RD (confidencialidad, custodia, titularidad)
for fila, name, alto in [(18, "Txt_Confid", 46), (19, "Txt_Custodia", 32), (20, "Txt_Titularidad", 32)]:
    j.merge_cells(f"B{fila}:H{fila}")
    c = j.cell(fila, 2, f'={ref(name)}'); c.font = Font(name=TNR, size=9); c.alignment = wrapj
    j.row_dimensions[fila].height = alto

j.merge_cells("B22:H22")
j["B22"] = ("FIRMA DE LA PERSONA QUE AUTORIZA LA DESCARGA Y DA LA INDICACIÓN PARA LA MODALIDAD DE "
            "ENTREGA DE LOS ARCHIVOS DE TRANSFERENCIA")
j["B22"].font = d_txt_b; j["B22"].alignment = wrapj
j.row_dimensions[22].height = 30
j.merge_cells("B23:H23")
j["B23"] = (f'={ref("Txt_En")}&{ref("Cfg_CiudadFirma")}&{ref("Txt_A")}&{fecha_txt("Dat_FechaInforme")}')
j["B23"].font = d_txt
j.merge_cells("B24:H24")
j["B24"] = (f'=IF({ref("Dat_TipoOperacion")}="Transferencia correcta","",'
            f'"— NO APLICA: esta intervención está marcada como Intransferibilidad —")')
j["B24"].font = Font(name=TNR, size=11, bold=True, color="C00000"); j["B24"].alignment = ctr
j.row_dimensions[25].height = 40
j.merge_cells("B26:D26"); j["B26"].border = firma_l
j.merge_cells("F26:H26"); j["F26"].border = firma_l
j["B27"] = "Persona que autoriza"; j["B27"].font = d_peq
j.merge_cells("F27:H27")
j["F27"] = f'="Técnico: "&{txt("Dat_Tecnico")}'; j["F27"].font = d_peq
j.merge_cells("F28:H28")
j["F28"] = f'={ref("Txt_Sello")}&{ref("Cfg_Empresa")}'; j["F28"].font = d_txt

j.merge_cells("B30:H30")
c = j.cell(30, 2,
    f'="Formato "&{ref("Cfg_CodJust")}&" · Versión "&{ref("Cfg_Version")}&" · Edición "&'
    f'{fecha_txt("Cfg_FechaEdicion")}&" · Elaborado: "&{ref("Cfg_Elaborado")}&" · Aprobado: "&'
    f'{ref("Cfg_Aprobado")}&" · UNE 66102:2025"')
c.font = d_peq; c.alignment = ctr; c.border = Border(top=Side(style="thin", color="000000"))

for col, w in zip("ABCDEFGH", [2, 22, 12, 11, 12, 16, 8, 10]):
    j.column_dimensions[col].width = w
setup_a4(j, "'JUSTIFICANTE TRANSFERENCIA'!$B$1:$H$30")
j.protection.sheet = True

# ============================================================ INTRANSF. CLIENTE
ic = wb.create_sheet("INTRANSFERIBILIDAD CLIENTE")
cab_centro(ic)
ic.merge_cells("A9:G9")
ic["A9"] = "Acuse de recibo Certificado de Intransferibilidad de datos"; ic["A9"].font = d_tit
ic.merge_cells("A11:G11")
ic["A11"] = f'={txt("Dat_Empresa")}'; ic["A11"].font = d_dato
ic["A14"] = "Estimados señores:"; ic["A14"].font = d_txt
ic.merge_cells("A16:G18")
ic["A16"] = ("En cumplimiento del requisito expresado en la disposición adicional primera, apartado 10, "
             "del Real decreto 125/2017, les remitimos copia del certificado de intransferibilidad "
             "correspondiente a la sustitución del tacógrafo:")
ic["A16"].font = d_txt; ic["A16"].alignment = wrapj

FICHA = [(20, "Modelo:", "Dat_Modelo", False), (21, "Nº de Serie:", "Dat_Serie", False),
         (23, "Montado en el vehículo:", "Dat_Matricula", True),
         (24, "Nº de informe/Certificado:", "Dat_NumInforme", False)]
for ws in ():
    pass
def ficha(ws):
    for fila, label, name, up in FICHA:
        ws.cell(fila, 1, label).font = d_txt
        ws.merge_cells(f"C{fila}:F{fila}")
        v = f'=UPPER({txt(name)})' if up else f'={txt(name)}'
        ws.cell(fila, 3, v).font = d_dato
    ws.cell(25, 1, "Fecha Informe:").font = d_txt
    c = ws.cell(25, 3, f'=IF({ref("Dat_FechaInforme")}="","",{ref("Dat_FechaInforme")})')
    c.font = d_dato; c.number_format = "DD/MM/YYYY"
ficha(ic)

ic.cell(27, 1, "Nombre:").font = d_txt
ic.merge_cells("C27:F27"); ic.cell(27, 3, f'={txt("Dat_NombreReceptor")}').font = d_dato
ic.cell(28, 1, "DNI:").font = d_txt
ic.merge_cells("C28:F28"); ic.cell(28, 3, f'={txt("Dat_DniReceptor")}').font = d_dato

ic.merge_cells("A30:G32")
ic["A30"] = (f'="En calidad de personal/propietario de la organización de transportes propietaria del '
             f'vehículo con matrícula "&UPPER({txt("Dat_Matricula")})&" declaro haber recibido el certificado '
             f'de intransferibilidad de fecha "&{fecha_txt("Dat_FechaInforme")}&" correspondiente a la '
             f'intervención técnica realizada sobre el vehículo de matrícula antedicha."')
ic["A30"].font = d_txt; ic["A30"].alignment = wrapj
ic.merge_cells("A34:G36")
ic["A34"] = ("En el caso de que el receptor no sea propietario de la organización de transportes, se "
             "compromete explícitamente, por el presente compromiso firmado, a entregar este documento "
             "(el certificado de intransferibilidad) a la propiedad de la citada organización")
ic["A34"].font = d_txt; ic["A34"].alignment = wrapj
ic.merge_cells("A38:G38")
ic["A38"] = f'=IF({ref("Dat_EntregaAparato")}="Sí",{ref("Txt_Entrega_Si")},"")'; ic["A38"].font = d_txt_b
ic.merge_cells("A39:G39")
ic["A39"] = f'=IF({ref("Dat_Achatarrar")}="Sí",{ref("Txt_Achatarrar_Si")},"")'; ic["A39"].font = d_txt_b
ic.merge_cells("A41:G42")
ic["A41"] = f'={ref("Txt_Confid")}'; ic["A41"].font = Font(name=TNR, size=9); ic["A41"].alignment = wrapj

ic["A44"] = "Entregado"; ic["A44"].font = d_txt
ic["A45"] = f'={ref("Cfg_CiudadFirma")}&" a"'; ic["A45"].font = d_txt
c = ic.cell(45, 3, f'=IF({ref("Dat_FechaEntrega")}="","",{ref("Dat_FechaEntrega")})')
c.font = d_dato; c.number_format = "DD/MM/YYYY"
ic["A47"] = "Firma:"; ic["A47"].font = d_txt
ic.merge_cells("B48:E48"); ic["B48"].border = firma_l
ic.merge_cells("A50:G50")
ic["A50"] = (f'=IF({ref("Dat_TipoOperacion")}="Intransferibilidad","",'
             f'"— NO APLICA: esta intervención está marcada como Transferencia correcta —")')
ic["A50"].font = Font(name=TNR, size=11, bold=True, color="C00000"); ic["A50"].alignment = ctr
pie_control(ic, 52, "Cfg_CodCliente")
for col, w in zip("ABCDEFG", [26, 10, 16, 12, 14, 14, 10]):
    ic.column_dimensions[col].width = w
setup_a4(ic, "'INTRANSFERIBILIDAD CLIENTE'!$A$1:$G$52")
ic.protection.sheet = True

# ============================================================ INTRANSF. ADMIN
ia = wb.create_sheet("INTRANSF. ADMINISTRACION")
cab_centro(ia)
ia.merge_cells("A9:G9")
ia["A9"] = "Acuse de recibo Certificado de Intransferibilidad de datos"; ia["A9"].font = d_tit
ia.merge_cells("A11:G11")
ia["A11"] = f'={ref("Cfg_Destinatario")}'; ia["A11"].font = d_dato
ia["A14"] = "Estimados señores:"; ia["A14"].font = d_txt
ia.merge_cells("A16:G18")
ia["A16"] = ("En cumplimiento del requisito expresado en la disposición adicional primera, apartado 10, "
             "del Real decreto 125/2017, les remitimos copia del certificado de intransferibilidad "
             "correspondiente a la sustitución del tacógrafo:")
ia["A16"].font = d_txt; ia["A16"].alignment = wrapj
ficha(ia)
ia["A28"] = "Entregado"; ia["A28"].font = d_txt
ia["A29"] = f'={ref("Cfg_CiudadFirma")}&" a"'; ia["A29"].font = d_txt
c = ia.cell(29, 3, f'=IF({ref("Dat_FechaEntrega")}="","",{ref("Dat_FechaEntrega")})')
c.font = d_dato; c.number_format = "DD/MM/YYYY"
ia.merge_cells("A32:G32")
ia["A32"] = (f'=IF({ref("Dat_TipoOperacion")}="Intransferibilidad","",'
             f'"— NO APLICA: esta intervención está marcada como Transferencia correcta —")')
ia["A32"].font = Font(name=TNR, size=11, bold=True, color="C00000"); ia["A32"].alignment = ctr
pie_control(ia, 34, "Cfg_CodAdmin")

ia["A38"] = "TEXTO PARA EL TRÁMITE TELEMÁTICO DE LA GENERALITAT (no se imprime)"
ia["A38"].font = f_seccion; ia["A38"].fill = fill_sec
for col in "BCDEFG": ia[f"{col}38"].fill = fill_sec
ia["A40"] = "Assumpte:"; ia["A40"].font = f_label
ia.merge_cells("C40:G40")
ia["C40"] = "Certificat de Intransferibilitat de dades Tacògraf digital"; ia["C40"].font = f_label
ia["A41"] = "Nom del fitxer:"; ia["A41"].font = f_label
ia.merge_cells("C41:G41")
ia["C41"] = f'={txt("Dat_NumInforme")}&{ref("Txt_Fitxer")}'; ia["C41"].font = f_label
ia["A43"] = "Exposo / Sol·licito:"; ia["A43"].font = f_label
ia["A44"] = "Copiar esta celda y pegarla en el formulario del trámite."; ia["A44"].font = f_nota
ia.merge_cells("A45:G49")
ia["A45"] = (f'={ref("Txt_Cat_1")}&{txt("Dat_Modelo")}&{ref("Txt_Cat_2")}&{txt("Dat_Serie")}&'
             f'{ref("Txt_Cat_3")}&UPPER({txt("Dat_Matricula")})&{ref("Txt_Cat_4")}&{txt("Dat_NumInforme")}&'
             f'{ref("Txt_Cat_5")}&{fecha_txt("Dat_FechaInforme")}&{ref("Txt_Cat_6")}')
ia["A45"].font = f_label; ia["A45"].alignment = wrap
ia["A45"].fill = fill_in; ia["A45"].border = box
ia["A51"] = "Enlace al trámite:"; ia["A51"].font = f_label
ia.merge_cells("C51:G51")
ia["C51"] = f'=HYPERLINK({ref("Cfg_UrlTramite")},"Petició genèrica — gencat.cat")'; ia["C51"].font = f_link
ia["A52"] = "Formulario directo (OVT):"; ia["A52"].font = f_label
ia.merge_cells("C52:G52")
ia["C52"] = f'=HYPERLINK({ref("Cfg_UrlTramiteOVT")},"Obrir el formulari a l\u2019OVT")'; ia["C52"].font = f_link
ia["A53"] = "Email del centro:"; ia["A53"].font = f_label
ia.merge_cells("C53:G53")
ia["C53"] = f'={ref("Cfg_Email")}'; ia["C53"].font = f_label
for col, w in zip("ABCDEFG", [26, 10, 16, 12, 14, 14, 10]):
    ia.column_dimensions[col].width = w
setup_a4(ia, "'INTRANSF. ADMINISTRACION'!$A$1:$G$34")
ia.protection.sheet = True

# ============================================================ REGISTRO
rg = wb.create_sheet("REGISTRO TRANSFERENCIAS")
rg.sheet_view.showGridLines = False
rg["A1"] = "REGISTRO DE TRANSFERENCIAS Y DE DESTRUCCIÓN DE ARCHIVOS"
rg["A1"].font = f_titulo
rg["A2"] = ("UNE 66102:2025, 7.5.1 d) y 8.5.1 — registro de transferencias realizadas y documento de "
            "destrucción al año. La fila 5 se rellena sola con la intervención en curso; cópiala como "
            "valores a la primera fila libre al cerrar cada expediente.")
rg["A2"].font = f_nota; rg["A2"].alignment = wrap
rg.row_dimensions[2].height = 30

CAB = ["Nº informe", "Fecha transferencia", "Matrícula", "Nº bastidor", "Nº serie UIV",
       "Empresa", "Técnico", "Modalidad entrega", "Fecha envío",
       "Destruir a partir de", "Fecha destrucción", "Método de destrucción",
       "Persona que destruye", "Firma digital del archivo (hash)"]
for i, t in enumerate(CAB, start=1):
    c = rg.cell(4, i, t); c.font = Font(name=ARIAL, size=9, bold=True)
    c.fill = fill_cab; c.border = box; c.alignment = ctrw
rg.row_dimensions[4].height = 34

AUTO = [f'={txt("Dat_NumInforme")}', f'=IF({ref("Dat_FechaTransf")}="","",{ref("Dat_FechaTransf")})',
        f'=UPPER({txt("Dat_Matricula")})', f'={txt("Dat_Bastidor")}', f'={txt("Dat_Serie")}',
        f'={txt("Dat_Empresa")}', f'={txt("Dat_Tecnico")}', f'={txt("Dat_Modalidad")}',
        f'=IF({ref("Dat_FechaEnvio")}="","",{ref("Dat_FechaEnvio")})',
        f'=IF({ref("Dat_FechaTransf")}="","",EDATE({ref("Dat_FechaTransf")},12))', "", "", "", ""]
for i, f in enumerate(AUTO, start=1):
    c = rg.cell(5, i, f if f else None)
    c.font = Font(name=ARIAL, size=9, bold=(i <= 10)); c.border = box
    if i in (2, 9, 10): c.number_format = "DD/MM/YYYY"
    if i > 10:
        c.fill = fill_in; c.protection = UNLOCK
rg.cell(5, 15, "← intervención en curso (automática)").font = f_nota

for fila in range(6, 40):
    for i in range(1, 15):
        c = rg.cell(fila, i); c.border = box; c.protection = UNLOCK
        c.font = Font(name=ARIAL, size=9)
        if i in (2, 9, 10, 11): c.number_format = "DD/MM/YYYY"
    rg.cell(fila, 10, f'=IF(B{fila}="","",EDATE(B{fila},12))').font = Font(name=ARIAL, size=9)
    rg.cell(fila, 10).protection = Protection(locked=True)
    rg.cell(fila, 15, f'=IF(B{fila}="","",IF(K{fila}<>"","Destruido",'
                      f'IF(TODAY()>EDATE(B{fila},12),"⚠ PENDIENTE DE DESTRUIR","En custodia")))'
            ).font = Font(name=ARIAL, size=9, bold=True)
rg.cell(4, 15, "Estado").font = Font(name=ARIAL, size=9, bold=True)
rg.cell(4, 15).fill = fill_cab; rg.cell(4, 15).border = box; rg.cell(4, 15).alignment = ctrw
rg.conditional_formatting.add("O6:O39", FormulaRule(formula=['$O6="⚠ PENDIENTE DE DESTRUIR"'], fill=fill_err))

for col, w in zip("ABCDEFGHIJKLMNO", [16, 13, 11, 18, 12, 24, 14, 14, 12, 13, 13, 20, 18, 26, 20]):
    rg.column_dimensions[col].width = w
rg.freeze_panes = "A5"
rg.page_setup.orientation = "landscape"
rg.page_setup.paperSize = rg.PAPERSIZE_A4
rg.page_setup.fitToWidth = 1
rg.page_setup.fitToHeight = 0
rg.sheet_properties.pageSetUpPr.fitToPage = True
rg.print_area = "'REGISTRO TRANSFERENCIAS'!$A$1:$O$39"
rg.print_title_rows = "4:4"
rg.protection.sheet = True

# ============================================================ cierre
for name, (sheet, coord) in nombres.items():
    wb.defined_names.add(DefinedName(name, attr_text=f"{quote_sheetname(sheet)}!{absolute_coordinate(coord)}"))
wb.move_sheet("CONFIGURACION", offset=6)
wb.save(OUT)
print("OK ->", OUT, "| hojas:", wb.sheetnames)
